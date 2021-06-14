"""Implement fiat model class"""

from ast import literal_eval
from hydromt.models.model_api import Model
from openpyxl import load_workbook
from os.path import abspath, basename, dirname, isfile, join
from pathlib import Path
from shapely.geometry import box
from shutil import copy
from typing import Dict, List, Optional, Set, Tuple
import geopandas as gpd
import glob
import hydromt
import logging
import numpy as np
import pandas as pd
import xarray as xr

from . import workflows, DATADIR

__all__ = ["FiatModel"]

logger = logging.getLogger(__name__)


class FiatModel(Model):
    """General and basic API for the FIAT model in hydroMT."""

    _NAME = "fiat"
    _CONF = "fiat_configuration.xlsx"
    _GEOMS = {}  # FIXME Mapping from hydromt names to model specific names
    _MAPS = {}  # FIXME Mapping from hydromt names to model specific names
    _FOLDERS = ["hazard", "exposure", "vulnerability", "output"]
    _DATADIR = DATADIR

    def __init__(
        self,
        root,
        mode="w",
        config_fn=None,
        data_libs=None,
        logger=logger,
        deltares_data=False,
        artifact_data=False,
    ):
        # Check if the root parameter exists.
        if not isinstance(root, (str, Path)):
            raise ValueError("The 'root' parameter should be a of str or Path.")

        super().__init__(
            root=root,
            mode=mode,
            config_fn=config_fn,
            data_libs=data_libs,
            deltares_data=deltares_data,
            artifact_data=artifact_data,
            logger=logger,
        )

    def setup_basemaps(
        self,
        region,
        **kwargs,
    ):
        """Define the model region.

        Adds model layer:

        * **region** geom: model region

        Parameters
        ----------
        region: dict
            Dictionary describing region of interest, e.g. {'bbox': [xmin, ymin, xmax, ymax]}.
            See :py:meth:`~hydromt.workflows.parse_region()` for all options.
        """

        kind, region = hydromt.workflows.parse_region(region, logger=self.logger)
        if kind == "bbox":
            geom = gpd.GeoDataFrame(geometry=[box(*region["bbox"])], crs=4326)
        elif kind == "grid":
            geom = region["grid"].raster.box
        elif kind == "geom":
            geom = region["geom"]
        else:
            raise ValueError(
                f"Unknown region kind {kind} for FIAT, expected one of ['bbox', 'grid', 'geom']."
            )

        # Set the model region geometry (to be accessed through the shortcut self.region).
        self.set_staticgeoms(geom, "region")

    def setup_hazard(
        self,
        map_fn,
        return_period=None,
        variable=None,
        **kwargs,
    ):
        """Add a hazard map to the FIAT model schematization.

        Adds model layer:

        * **hazard** map(s): A hazard map with the nomenclature <hazard_type>_rp<return_period>.

        Parameters
        ----------
        map_fn: (list of) str, Path
            Path to hazard raster file.
        return_period: (list of) int, float, optional
            Return period in years, required for a risk calculation. The default is None.
        variable: str
            Name of variable in dataset, required when reading NetCDF data.
        """

        # Check the input parameters.
        rp_lst, var_lst = [], []
        fn_lst = [map_fn] if isinstance(map_fn, (str, Path)) else map_fn
        if not (isinstance(fn_lst, list) and isinstance(fn_lst[0], (str, Path))):
            raise TypeError(f"The 'map_fn' parameter should be a of str or Path.")
        if return_period is not None:
            rp_lst = (
                [return_period]
                if isinstance(return_period, (int, float))
                else return_period
            )
            if not isinstance(rp_lst, list):
                raise TypeError(
                    f"The 'return_period' parameter should be a float or int, received a {type(return_period)} instead."
                )
            if not (len(rp_lst) == len(fn_lst) or len(rp_lst) == len(var_lst)):
                raise IndexError(
                    "The length of 'return_period' parameter should match with the 'map_fn' or 'variable' parameters."
                )
        if variable is not None:
            var_lst = [variable] if isinstance(variable, str) else variable
            if not isinstance(var_lst, list):
                raise TypeError(
                    f"The 'variable' parameter should be a str, received a {type(variable)} instead."
                )
            if len(fn_lst) == 1:
                kwargs.update(variables=var_lst)

        # Get a (clipped) Dataset of hazard maps (also for a single hazard map input).
        rp = None
        for i, fn in enumerate(fn_lst):
            if str(fn).endswith(".nc"):
                if variable is None:
                    raise ValueError(
                        "The 'variable' parameter is required when reading netcdf data."
                    )
                kwargs.update(driver="netcdf")
            if len(fn_lst) == len(var_lst):
                kwargs.update(variables=[var_lst[i]])
            if len(fn_lst) == len(rp_lst):
                rp = rp_lst[i]
            ds = self.data_catalog.get_rasterdataset(
                fn, geom=self.region, single_var_as_array=False, **kwargs
            )
            if self.staticmaps and not self.staticmaps.raster.identical_grid(ds):
                raise ValueError("The hazard maps should have identical grids.")

            # Correct (if necessary) the grid orientation from the lower to the upper left corner.
            if ds.raster.res[1] > 0:
                ds = ds.reindex({ds.raster.y_dim: list(reversed(ds.raster.ycoords))})

            # Rename the hazard map and add to staticmaps.
            hazard_type = self.get_config("hazard_type", fallback="flooding")
            for j, name in enumerate(ds.data_vars):
                da = ds[name]

                # Get the return period per input parameter.
                if len(rp_lst) == len(var_lst) == len(ds.data_vars):
                    rp = rp_lst[j]

                # Get (if possible) the return period from dataset names if the input parameter is None.
                elif len(ds.data_vars) > 1:
                    try:
                        # Get the numeric part (including ".") after "rp" (if in name).
                        fstrip = lambda x: x in "0123456789."
                        rps = "".join(filter(fstrip, name.lower().split("rp")[-1]))
                        rp = literal_eval(rps)
                        assert isinstance(rp, (int, float))
                    except (ValueError, AssertionError):
                        raise ValueError(
                            f"Could not derive the return periods for hazard map: {name}."
                        )

                # Rename the hazard file name.
                rp_str = f"rp{rp}" if rp is not None else ""
                out_name = f"{hazard_type}_{rp_str}"
                hazard_fn = join(
                    self.root, "hazard", f"{out_name}.tif"
                )  # TODO: Comment to disable absolute paths!
                # hazard_fn = join("hazard", f"{out_name}.tif") # TODO: Uncomment to enable relative paths with respect to root!

                # Add the hazard map to staticmaps.
                self.set_staticmaps(da, out_name)

                # Update the config with filepath.
                if rp is not None:
                    post = f"(rp {rp})"
                    self.set_config("hazard", rp, hazard_fn)
                else:
                    post = ""
                    self.set_config("hazard", hazard_fn)
                self.logger.debug(
                    f"Added {hazard_type} hazard layer: {out_name} {post}"
                )

    def setup_exposure_buildings(
        self,
        bld_fn=None,
        pop_fn=None,
        country=None,
        unit="USD",
    ):
        """Add a buildings value exposure map to the FIAT model schematization.

        Adds model layer:

        * **exposure** map: An exposure map with the nomenclature <exposure_type>.

        Parameters
        ----------
        bld_fn: str
            Nametag of the global building footprints dataset (raster file).
        pop_fn: str
            Nametag of the global population dataset (raster file).
        country: str, optional
            Country name or tag in which the hazard region is location. Required to derive the damage function and
            potential damage values associated to the hazard region.
        """

        if bld_fn and pop_fn:
            # Clip the building footprint map from the global dataset and store as a xarray.DataArray.
            da_bld = self.data_catalog.get_rasterdataset(
                bld_fn,
                geom=self.region,
                buffer=4,
            ).rename("bld")

            # Clip the population map from the global dataset and store as a xarray.DataArray.
            da_pop = self.data_catalog.get_rasterdataset(
                pop_fn,
                geom=self.region,
                buffer=4,
            ).rename("pop")

            # Create the population, buildings and population per building count maps and store as a xarray.DataSet.
            ds_count = workflows.create_population_per_building_map(
                da_bld,
                da_pop,
                ds_like=self.staticmaps,
                logger=self.logger,
            )

            # Add the exposure count maps to staticmaps.
            self.set_staticmaps(ds_count["bld"], name="buildings_count")
            self.set_staticmaps(ds_count["pop"], name="population_count")
            self.set_staticmaps(ds_count["pop_bld"], name="population_buildings_count")

        # Create the buildings value map.
        if "population_buildings_count" in self.staticmaps.data_vars:
            # Get the associated country tag (alpha-3 code).
            tag = self._get_country_tag(country)

            # Get the associated vulnerability (damage function id and maximum damage value).
            df_id, max_damage = self._get_vulnerability()
            self.logger.debug(
                "Calculating building values with maximum damage: "
                f"{max_damage:.2f} {unit:s}/person (country = {tag:s})."
            )

            # Create a building value map.
            ds_bld_value = self.staticmaps["population_buildings_count"] * max_damage
            ds_bld_value.raster.set_nodata(nodata=0)
            ds_bld_value.name = "bld_value"

            # Add the buildings value map to staticmaps.
            self.set_staticmaps(ds_bld_value, name="buildings_value")

        # Check if the buildings value map has correctly been generated.
        else:
            raise ValueError(
                "The buildings value exposure layer is not correctly generated."
            )

        # Add the buildings value map to the config.
        exp_layer = {
            "use": 1,
            "category": "Buildings Value",
            "max_damage": 1,
            "bu": 0,
            "function": df_id,
            "map": 1,
            "scale_factor": 1,
            "raster": join(
                self.get_config("exposure", fallback=""), "buildings_value.tif"
            ),
            "unit": unit,
        }

        lyrs = [int(key.split("_")[1]) for key in self.config if key.startswith("exp_")]
        if lyrs:
            for i in lyrs:
                if self.config[f"exp_{i}"]["category"] == "Buildings Value":
                    lyr = int(i)
                else:
                    lyr = max(lyrs) + 1
        else:
            lyr = 1

        self.set_config(f"exp_{lyr}", exp_layer)
        self.logger.debug(f"Added exposure layer: Buildings Value")

    def setup_exposure_roads(
        self,
        exposure_fn,
        category,
        unit,
    ):
        # TODO general method to set exposure layer
        pass

    def scale_exposure(
        self,
        scenario,
        year,
    ):
        """Scale the exposure to the forecast year, using the shared socioeconomic pathway (SSP) projections for
        population and GDP growth.

        Parameters
        ----------
        scenario: str
            Nametag of the shared socioeconomic pathway (SSP), required for a forecast calculation.
        year: int
            The forecast year to which the exposure data is scaled.
        """

        # Set the scenario and year config parameters.
        self.set_config("scenario", scenario)
        self.set_config("year", year)

        # Determine the scale factor.
        pop_correction = self._get_population_correction_factor()
        gdp_correction = self._get_gdp_correction_factor()
        scale_factor = pop_correction * gdp_correction

        # Set the scale factor.
        for row in range(7, 99):
            exp_layer = self.get_config(f"exp_{row - 6}", fallback={})
            if not exp_layer:
                break

            exp_layer["scale_factor"] = scale_factor
            self.set_config(f"exp_{row - 6}", exp_layer)

    def _get_country_tag(self, country):
        """Return the country tag for a country name input."""

        # Get the country tag from the country name.
        if country or "country" in self.config:
            if not country:
                country = self.config["country"]

            # Read the global exposure configuration.
            df_config = pd.read_excel(
                Path(self._DATADIR).joinpath("global_configuration.xlsx"),
                sheet_name="Buildings",
            )

            # Extract the country tag.
            if len(country) > 3:
                tag = (
                    df_config.loc[
                        df_config["Country_Name"] == country, "Alpha-3"
                    ].values[0]
                    if country in df_config["Country_Name"].tolist()
                    else None
                )
            else:
                tag = country

            # If the country tag is not valid, get the country tag from nearest country.
            if not tag in df_config["Alpha-3"].tolist():
                tag = self._get_nearest_country()
                self.logger.debug(
                    "The country tag (related to the country name) is not valid."
                    "The country tag of the nearest country is used instead."
                )

        else:
            # If the country parameter is None, get the country tag from nearest country.
            country = self._get_nearest_country()

        # Set the country tag.
        self.set_config("country", tag)
        return tag

    def _get_nearest_country(self):
        """Return the country tag of the nearest country."""

        # Read the global exposure configuration.
        df_config = pd.read_excel(
            Path(self._DATADIR).joinpath("global_configuration.xlsx"),
            sheet_name="Buildings",
        )

        # TODO: Lookup country from shapefile!
        pass

    def _get_vulnerability(self):
        """Return the damage function id and the maximum damage number."""

        # Read the global exposure configuration.
        df_config = pd.read_excel(
            Path(self._DATADIR).joinpath("global_configuration.xlsx"),
            sheet_name="Buildings",
        )

        # Get the damage function id.
        df_id = df_config.loc[
            df_config["Alpha-3"] == self.config["country"],
            f"Damage_Function_ID_{self.config['hazard_type'].capitalize()}",
        ].values[0]

        # Get the maximum damage value.
        max_damage = df_config.loc[
            df_config["Alpha-3"] == self.config["country"],
            f"Max_Damage_{self.config['hazard_type']. capitalize()}",
        ].values[0]

        return df_id, max_damage

    def _get_population_correction_factor(self):
        """ """

        # Read the global SSP data.
        df_pop = pd.read_excel(
            Path(self._DATADIR).joinpath("growth_scenarios", "global_pop.xlsx"),
            sheet_name="Data",
        )

        # Extract the national data.
        pop_data = (
            df_pop.loc[
                (df_pop["Region"] == self.config["country"])
                & (df_pop["Scenario"] == self.config["scenario"])
            ]
            .reset_index(drop=True)
            .iloc[:, 5:-1]
        )

        # In case multiple data sources are available, use the averaged values.
        pop_data = pop_data.mean(axis=0)

        # Interpolate (linear) the data to obtain annual results.
        annual_pop_data = np.array(
            range(int(pop_data.index[0]), int(pop_data.index[-1]) + 1, 1)
        )
        interp_pop_data = list(
            np.interp(
                annual_pop_data,
                pop_data.index.astype(int).values,
                pop_data.values.astype(float),
            )
        )

        # Determine the indexes of the reference year (GHS 2015) and the forecast year.
        ref_year_idx = list(annual_pop_data).index(2015)
        forecast_year_idx = list(annual_pop_data).index(self.config["year"])

        # Calculate the correction factor.
        pop_correction = (
            interp_pop_data[forecast_year_idx] / interp_pop_data[ref_year_idx]
        )

        return pop_correction

    def _get_gdp_correction_factor(self):
        """ """

        # Read the global SSP data.
        df_pop = pd.read_excel(
            Path(self._DATADIR).joinpath("growth_scenarios", "global_pop.xlsx"),
            sheet_name="Data",
        )
        df_gdp = pd.read_excel(
            Path(self._DATADIR).joinpath("growth_scenarios", "global_gdp(ppp).xlsx"),
            sheet_name="Data",
        )

        # Extract the national data.
        pop_data = (
            df_pop.loc[
                (df_pop["Region"] == self.config["country"])
                & (df_pop["Scenario"] == self.config["scenario"])
            ]
            .reset_index(drop=True)
            .iloc[:, 5:-1]
        )
        gdp_data = (
            df_gdp.loc[
                (df_gdp["Region"] == self.config["country"])
                & (df_gdp["Scenario"] == self.config["scenario"])
            ]
            .reset_index(drop=True)
            .iloc[:, 5:-1]
        )

        # In case multiple data sources are available, use the averaged values.
        pop_data = pop_data.mean(axis=0)
        gdp_data = gdp_data.mean(axis=0)

        # Determine the GDP(PPP) per capita and interpolate (linear) the data to obtain annual results.
        gdp_ppp_data = gdp_data * 1000 / pop_data
        annual_gdp_per_cap_data = np.array(
            range(int(gdp_ppp_data.index[0]), int(gdp_ppp_data.index[-1]) + 1, 1)
        )
        interp_gdp_per_cap_data = list(
            np.interp(
                annual_gdp_per_cap_data,
                gdp_ppp_data.index.astype(int).values,
                gdp_ppp_data.values.astype(float),
            )
        )

        # FIXME: why is the reference factor fixe to 2019?!
        # Determine the indexes of the reference year (2019) and the forecast year.
        ref_year_idx = list(annual_gdp_per_cap_data).index(2019)
        forecast_year_idx = list(annual_gdp_per_cap_data).index(self.config["year"])

        # Calculate the correction factor.
        correction_factor = (
            annual_gdp_per_cap_data[forecast_year_idx]
            / annual_gdp_per_cap_data[ref_year_idx]
        )

        return correction_factor

    # Overwrite the model_api methods for root and config.
    def set_root(self, root=None, mode="w"):
        """Initialized the model root.
        In read mode it checks if the root exists.
        In write mode in creates the required model folder structure.

        Parameters
        ----------
        root: str, optional
            Path to model root.
        mode: {"r", "r+", "w"}, optional
            Read/write-only mode for model files.
        """
        # Do super method and update absolute paths in config.
        if root is None and self.root is not None:
            root = self.root
        super().set_root(root=root, mode=mode)
        if self._write and root is not None:
            root = abspath(root)

            # Set the general information.
            # TODO: Comment to disable absolute paths!
            self.set_config("vulnerability", join(root, "vulnerability"))
            self.set_config("exposure", join(root, "exposure"))
            self.set_config("output", join(root, "output"))
            # TODO: Uncomment to enable relative paths with respect to the root!
            # self.set_config("vulnerability", "vulnerability")
            # self.set_config("exposure", "exposure")
            # self.set_config("output", "output")
            # Set the hazard information.
            hazard_maps = self.get_config("hazard")
            if isinstance(hazard_maps, dict):
                # TODO: Comment to disable absolute paths!
                hazard_maps = {
                    k: join(root, "hazard", basename(v)) for k, v in hazard_maps.items()
                }
                # TODO: Uncomment to enable relative paths with respect to the root!
                # hazard_maps = {k: join("hazard", basename(v)) for k, v in hazard_maps.items()}
                self.set_config("hazard", hazard_maps)
            elif isinstance(hazard_maps, str):
                # TODO: Comment to disable absolute paths!
                self.set_config("hazard", join(root, "hazard", basename(hazard_maps)))
                # TODO: Uncomment to enable relative paths with respect to the root!
                # self.set_config("hazard", join("hazard", basename(hazard_maps)))

            # Set the exposure information.
            for row in range(7, 99):
                exp_layer = self.get_config(f"exp_{row - 6}", fallback={})
                if not exp_layer:
                    break

                exp_layer["raster"] = join(
                    self.config["exposure"], basename(exp_layer["raster"])
                )
                self.set_config(f"exp_{row - 6}", exp_layer)

    def _configread(self, fn):
        """Parse fiat_configuration.xlsx and risk.csv to dict."""

        wb = load_workbook(filename=fn)
        ws = wb["Input"]
        config = dict()

        # Read the general information.
        general_config = {
            "country": ws["B1"].value,
            "case": ws["B2"].value,
            "hazard_type": ws["B3"].value,
            "unit": ws["B5"].value,
            "scenario": ws["J1"].value,
            "year": ws["J2"].value,
            "vulnerability": ws["J3"].value,
            "exposure": ws["J4"].value,
            "output": ws["J5"].value,
        }
        general_config = {k: v for k, v in general_config.items() if v is not None}
        config.update(general_config)

        # Read the hazard information.
        hazard_fn = ws["B4"].value
        if hazard_fn:
            if Path(hazard_fn).suffix == ".csv":
                hazard_maps = {}
                with open(join(dirname(fn), "risk.csv"), "r") as f:
                    lines = f.readlines()
                    for line in lines[1:]:
                        map_fn, rp = line.strip("\n").split(",", maxsplit=2)

                        # Strip white spaces and parse rp to int or float with literal_eval.
                        hazard_maps.update({literal_eval(rp.strip()): map_fn.strip()})

                config.update({"hazard": hazard_maps})
            else:
                config.update({"hazard": hazard_fn})

        # Read the exposure information.
        for row in range(7, 99):
            if ws[f"A{row}"].value is None:
                break
            exp_layer = {
                "use": ws[f"A{row}"].value,
                "category": ws[f"B{row}"].value,
                "max_damage": ws[f"C{row}"].value,
                "bu": ws[f"D{row}"].value,
                "function": ws[f"F{row}"].value,
                "map": ws[f"H{row}"].value,
                "scale_factor": ws[f"I{row}"].value,
                "raster": ws[f"J{row}"].value,
                "unit": ws[f"K{row}"].value,
                "landuse": ws[f"L{row}"].value,
            }
            exp_layer = {k: v for k, v in exp_layer.items() if v is not None}
            config.update({f"exp_{row-6}": exp_layer})

        wb.close()
        return config

    def _configwrite(self, fn):
        """Write config to fiat_configuration.xlsx and risk.csv"""

        # Load the template.
        fn_temp = join(self._DATADIR, self._NAME, self._CONF)
        wb = load_workbook(filename=fn_temp)
        ws = wb["Input"]

        # Store the general information.
        general_config = {
            "B1": "country",
            "B2": "case",
            "B3": "hazard_type",
            "B5": "unit",
            "J1": "scenario",
            "J2": "year",
            "J3": "vulnerability",
            "J4": "exposure",
            "J5": "output",
        }
        for loc, key in general_config.items():
            value = self.get_config(key)
            if value is not None:
                ws[loc] = str(value)

        # Store the hazard information.
        hazard_maps = self.get_config("hazard")
        if isinstance(hazard_maps, dict):
            rpmin, rpmax = min(hazard_maps.keys()), max(hazard_maps.keys())
            with open(join(dirname(fn), "risk.csv"), "w") as f:
                f.write(f"{rpmin}, {rpmax}\n")
                for rp in sorted(hazard_maps.keys()):
                    # TODO: Comment to disable absolute paths!
                    map_fn = abspath(hazard_maps[rp])
                    # TODO: Uncomment to enable relative paths with respect to the root!
                    # map_fn = hazard_maps[rp]
                    f.write(f"{map_fn:s}, {rp}\n")

            # TODO: Comment to disable absolute paths!
            ws["B4"] = join(self.root, "risk.csv")
            # TODO: Uncomment to enable relative paths with respect to the root!
            # ws["B4"] = "risk.csv"
        elif isinstance(hazard_maps, str):
            ws["B4"] = abspath(hazard_maps)  # TODO: Comment to disable absolute paths!
            # ws["B4"] = hazard_maps  # TODO: Uncomment to enable relative paths with respect to the root!

        # Store the exposure information.
        for row in range(7, 99):
            exp_layer = self.get_config(f"exp_{row-6}", fallback={})
            if not exp_layer:
                break

            # Write with fallback options.
            ws[f"A{row}"] = int(exp_layer.get("use", 0))
            ws[f"B{row}"] = str(exp_layer.get("category", ""))
            ws[f"C{row}"] = float(exp_layer.get("max_damage", 0))
            ws[f"D{row}"] = float(exp_layer.get("bu", 0))
            ws[f"F{row}"] = str(exp_layer.get("function", ""))
            ws[f"H{row}"] = int(exp_layer.get("map", 1))
            ws[f"I{row}"] = float(exp_layer.get("scale_factor", 1))
            ws[f"J{row}"] = str(exp_layer.get("raster", ""))
            ws[f"K{row}"] = str(exp_layer.get("unit", ""))
            ws[f"L{row}"] = str(exp_layer.get("landuse", ""))

            # Copy the damage function file.

            # TODO: Comment to disable absolute paths!
            copy(
                Path(self._DATADIR).joinpath(
                    "damage_functions",
                    self.config["hazard_type"],
                    self.config["unit"],
                    f"{exp_layer.get('function', '')}.csv",
                ),
                Path(self.config["vulnerability"]).joinpath(
                    f"{exp_layer.get('function', '')}.csv"
                ),
            )

            # TODO: Uncomment to enable relative paths with respect to the root!
            # copy(
            #     Path(self._DATADIR).joinpath(
            #         "damage_functions", self.config["hazard_type"], self.config["unit"],
            #         f"{exp_layer.get('function', '')}.csv"
            #     ),
            #     Path(self.root).joinpath(self.config["vulnerability"], f"{exp_layer.get('function', '')}.csv")
            # )

        # Write to file.
        wb.save(fn)
        wb.close()

    ## I/O
    def read(self):
        """Method to read the complete model schematization and configuration from file."""

        self.logger.info(f"Reading model data from {self.root}")
        self.read_config()
        self.read_staticmaps()
        self.read_staticgeoms()

    def write(self):
        """Method to write the complete model schematization and configuration to file."""

        self.logger.info(f"Writing model data to {self.root}")
        # if in r, r+ mode, only write updated components
        if not self._write:
            self.logger.warning("Cannot write in read-only mode")
            return
        if self.config:  # try to read default if not yet set
            self.write_config()
        if self._staticmaps:
            self.write_staticmaps()
        if self._staticgeoms:
            self.write_staticgeoms()
        if self._forcing:
            self.write_forcing()

    def read_staticmaps(self):
        """Read staticmaps at <root/?/> and parse to xarray Dataset."""

        # to read gdal raster files use: hydromt.open_mfraster()
        # to read netcdf use: xarray.open_dataset()
        if not self._write:
            # start fresh in read-only mode
            self._staticmaps = xr.Dataset()

        # Read the hazard maps.
        hazard_maps = self.get_config("hazard")
        fn_lst = []
        if isinstance(hazard_maps, dict):
            fn_lst = list(hazard_maps.values())
        elif isinstance(hazard_maps, str):
            fn_lst = [hazard_maps]

        for fn in fn_lst:
            # fn = join(self.root, fn)  # TODO: Uncomment to enable relative paths with respect to the root!
            name = basename(fn).rsplit(".", maxsplit=2)[0]
            if not isfile(fn):
                logger.warning(f"Could not find hazard map at {fn}.")
            else:
                self.set_staticmaps(hydromt.open_raster(fn), name=name)

        # Read the exposure maps.
        exp_root = self.get_config("exposure")
        if exp_root is not None:
            # TODO: Comment to disable absolute paths!
            fns = glob.glob(join(exp_root, "*.tif"))
            # TODO: Uncomment to enable relative paths with respect to the root!
            # fns = glob.glob(join(self.root, exp_root, "*.tif"))
            if len(fns) == 0:
                logger.warning(f"Could not find any expsure maps in {exp_root}.")
            else:
                self.set_staticmaps(hydromt.open_mfraster(fns))

    def write_staticmaps(self, compress="lzw"):
        """Write staticmaps at <root/?/> in model ready format."""

        # to write to gdal raster files use: self.staticmaps.raster.to_mapstack()
        # to write to netcdf use: self.staticmaps.to_netcdf()
        if not self._write:
            raise IOError("Model opened in read-only mode.")
        hazard_type = self.get_config("hazard_type", fallback="flooding")
        hazard_maps = [n for n in self.staticmaps.data_vars if hazard_type in n]
        if len(hazard_maps) > 0:
            self.staticmaps[hazard_maps].raster.to_mapstack(
                join(self.root, "hazard"), compress=compress
            )
        exposure_maps = [n for n in self.staticmaps.data_vars if hazard_type not in n]
        if len(exposure_maps) > 0:
            exp_root = self.get_config("exposure")
            self.staticmaps[exposure_maps].raster.to_mapstack(
                exp_root, compress=compress
            )  # TODO: Comment to disable absolute paths!
            # TODO: Uncomment to enable relative paths with respect to the root!
            # self.staticmaps[exposure_maps].raster.to_mapstack(join(self.root, exp_root), compress=compress)

    def read_staticgeoms(self):
        """Read staticgeoms at <root/?/> and parse to dict of GeoPandas."""

        if not self._write:
            # start fresh in read-only mode
            self._staticgeoms = dict()
        region_fn = join(self.root, "region.GeoJSON")
        if isfile(region_fn):
            self.set_staticgeoms(gpd.read_file(region_fn), "region")
        return self._staticgeoms

    def write_staticgeoms(self):
        """Write staticmaps at <root/?/> in model ready format."""

        if not self._write:
            raise IOError("Model opened in read-only mode")
        if self._staticgeoms:
            for name, gdf in self._staticgeoms.items():
                gdf.to_file(join(self.root, f"{name}.geojson"), driver="GeoJSON")

    def read_forcing(self):
        """Read forcing at <root/?/> and parse to dict of xr.DataArray."""

        return self._forcing
        # raise NotImplementedError()

    def write_forcing(self):
        """write forcing at <root/?/> in model ready format."""
        pass
        # raise NotImplementedError()

    def read_states(self):
        """Read states at <root/?/> and parse to dict of xr.DataArray."""

        return self._states
        # raise NotImplementedError()

    def write_states(self):
        """write states at <root/?/> in model ready format."""

        pass
        # raise NotImplementedError()

    def read_results(self):
        """Read results at <root/?/> and parse to dict of xr.DataArray."""

        return self._results
        # raise NotImplementedError()

    def write_results(self):
        """write results at <root/?/> in model ready format."""

        pass
        # raise NotImplementedError()
